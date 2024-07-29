!pip install -qU langchain-openai #在安静模式（q）下升级（U）安装，用于与OpenAI的API进行交互
!pip install xlsxwriter #用于创建Excel文件的Python模块
!pip install pandas matplotlib openpyxl #openpyxl 是一个用于读写Excel文件；pandas用于数据处理和分析；matplotlib允许生成各种类型的图表和可视化；
!pip install tencentcloud-sdk-python #腾讯云官方SDK，用于API
!pip install --upgrade zhipuai


%pip install unbabel-comet
!huggingface-cli login

#############################################################
from itertools import cycle #用于创建一个无限循环的迭代器
from langchain_openai import ChatOpenAI
from langchain_core.prompts import ChatPromptTemplate #用于创建聊天提示模板
from langchain_core.pydantic_v1 import BaseModel, Field, conint #BaseModel用于创建数据模型，Field用于定义模型字段及其验证，conint 是一个类型约束，限制整数必须满足特定条件（例如，范围）。
from typing_extensions import Literal #用于类型注解中的文字类型
from typing import List, Optional #用于类型注解
import os #用于与操作系统进行交互，例如文件和目录的操作
import hashlib #用于进行各种安全哈希和消息摘要操作
import requests #用于发送HTTP请求，用于与网络服务进行交互
import json #用于解析和生成JSON格式的数据
from tencentcloud.common import credential #用于管理腾讯云的认证信息
from tencentcloud.common.profile.client_profile import ClientProfile #用于配置腾讯云客户端的基本属性
from tencentcloud.common.profile.http_profile import HttpProfile #用于配置HTTP相关属性，如请求超时时间
from tencentcloud.common.exception.tencent_cloud_sdk_exception import TencentCloudSDKException #用于处理腾讯云SDK异常
from tencentcloud.tmt.v20180321 import tmt_client, models #tmt_client用于创建腾讯云机器翻译客户端；model包含各种请求和响应模型
import time #用于时间相关操作，例如延迟和获取当前时间
import pandas as pd
import matplotlib.pyplot as plt #用于创建各种类型的图表和可视化
import matplotlib.colors as mcolors #用于处理颜色相关的操作
import colorsys #用于处理颜色空间转换，如RGB转换为HSV
from openpyxl import load_workbook #用于加载现有的Excel
from openpyxl.drawing.image import Image #用于在Excel中插入图像
from io import BytesIO #用于在内存中处理二进制数据流
from concurrent.futures import ThreadPoolExecutor, as_completed #用于并发执行任务和管理线程池
import datetime
import base64
import hmac
from zhipuai import ZhipuAI
from openai import OpenAI
import os


from openpyxl.utils.dataframe import dataframe_to_rows
from comet import download_model, load_from_checkpoint  

# 下载并加载翻译质量评估模型
model_path = download_model("Unbabel/wmt22-comet-da")
model_with_ref = load_from_checkpoint(model_path)

# 下载并加载翻译质量评估模型
model_path = download_model("Unbabel/wmt22-cometkiwi-da")
model_no_ref = load_from_checkpoint(model_path)


#############################################################

#多个API列表，用于OpenAI、Youdao、Baidu和Tencent的API调用。每个列表包含多个API密钥，以便在多个请求之间进行轮换使用，避免单个密钥的频繁调用导致限制
openai_api_keys = [
    "",  #替换为自己的openai_api_key
    ""   #替换为自己的openai_api_key
]

youdao_api_keys = [
    "", #替换为自己的youdao_app_key, youdao_app_secret
    "", #替换为自己的youdao_app_key, youdao_app_secret
    "", #替换为自己的youdao_app_key, youdao_app_secret
    ""  #替换为自己的youdao_app_key, youdao_app_secret
]

baidu_api_keys = [
    "", #替换为自己的baidu_appid, baidu_api_key
    "", #替换为自己的baidu_appid, baidu_api_key
    "", #替换为自己的baidu_appid, baidu_api_key
    ""  #替换为自己的baidu_appid, baidu_api_key
]

tencent_api_keys = [
    "", #替换为自己的tencent_secret_id, tencent_secret_key
    "", #替换为自己的tencent_secret_id, tencent_secret_key
    "", #替换为自己的tencent_secret_id, tencent_secret_key
    ""  #替换为自己的tencent_secret_id, tencent_secret_key
]

xfyun_api_keys = [
    "", #替换为自己的xfyun_app_id, xfyun_api_key, xfyun_api_secret
    "", #替换为自己的xfyun_app_id, xfyun_api_key, xfyun_api_secret
    "", #替换为自己的xfyun_app_id, xfyun_api_key, xfyun_api_secret
    ""  #替换为自己的xfyun_app_id, xfyun_api_key, xfyun_api_secret
]

zhipu_api_keys = [
    "", #替换为自己的zhipu_api_key
    "", #替换为自己的zhipu_api_key
    "", #替换为自己的zhipu_api_key
    ""  #替换为自己的zhipu_api_key
]

#创建一个空列表用于记录后续使用的OpenAI key
used_openai_keys = []
#使用cycle函数实现API的轮询，以避免单个密钥的频繁调用
openai_key_cycle = cycle(openai_api_keys)
youdao_key_cycle = cycle(youdao_api_keys)
baidu_key_cycle = cycle(baidu_api_keys)
tencent_key_cycle = cycle(tencent_api_keys)
xfyun_key_cycle = cycle(xfyun_api_keys)
zhipu_key_cycle = cycle(zhipu_api_keys)
#设置环境变量
os.environ["OPENAI_BASE_URL"] = "https://gtapi.xiaoerchaoren.com:8932/v1"

#定义函数，用于获取下一个API密钥
def get_next_openai_key():
    next_key = next(openai_key_cycle)
    os.environ["OPENAI_API_KEY"] = next_key #将该key设置为环境变量
    return next_key

def get_next_youdao_key():
    return next(youdao_key_cycle)

def get_next_baidu_key():
    return next(baidu_key_cycle)

def get_next_tencent_key():
    return next(tencent_key_cycle)

def get_next_xfyun_key():
    return next(xfyun_key_cycle)

def get_next_zhipu_key():
    return next(zhipu_key_cycle)

# 初始化ChatOpenAI对象
def initialize_llm():
    api_key = get_next_openai_key()
    return ChatOpenAI(model="gpt-4o", api_key=api_key)

llm = initialize_llm()

# 创建ChatPromptTemplate
system_message = """
你是一个严谨的翻译专家，你接下来准备用MQM来对译文质量进行评估，评估时可以对照参考译文。你要准确地寻找出译文中的全部翻译错误。
JSON格式的MQM错误类型表如下：
'Mistranslation', 'Omission', 'Addition', 'Untranslated', 'Over-translation', 'Under-translation','Do Not Translate',
'Specific Culture Reference',
'Missing Text', 'Character Formatting',
'Grammar', 'Spelling', 'Punctuation', 'Unintelligible ', 'Character Encoding ',
'Date Format', 'Currency Format', 'Number Format', 'Measurement Format',
'Inconsistent Style', 'Inappropriate Style', 'Awkward Style', 'Organizational Style', 'Register', 'Unidiomatic Style',
'Inconsistent Terminology', 'Incorrect Terminology', 'Wrong Term'     
请你对照上面的MQM错误类型，找出译文中的全部翻译错误，使用MQM为错误归类；每个错误里提供详细解释，并给出正确的参考译文；并为整个句子的翻译质量打分，分数区间为0到1。
"""
prompt = ChatPromptTemplate.from_messages([
    ("system", system_message),
    ("user", "{input}"),
])


class TranslationError(BaseModel):
    source_chunk: str = Field(description="The source text chunk")
    target_chunk: str = Field(description="The target translation chunk")
    reference_chunk: str = Field(description="The reference chunk")
    error_type: str = Field(description="The type of translation error")
    severity: conint(ge=0, le=10) = Field(description="The severity of the error, from 0 (least severe) to 10 (most severe)")
    explanation: str = Field(description="A detailed explanation of the error with a correction suggestion")

class TranslationErrorReport(BaseModel):
    errors: List[TranslationError] = Field(description="List of translation errors")
    scores: conint(ge=0, le=1) = Field(description="The score of the translation, from 0 (poor quality) to 1 (perfect translation)")

tools = [TranslationErrorReport]
llm_with_tools = llm.bind_tools(tools, tool_choice="TranslationErrorReport")

# 组合提示模板和结构化语言模型
chain = prompt | llm_with_tools

#######################################################################

# Step 1: 读取数据集并解析txt文件
def read_txt(file_path):
    with open(file_path, 'r', encoding='utf-8') as file: #使用with语句确保文件在使用后自动关闭，即使在处理中发生错误也会如此
        data = file.readlines() #读取文件的所有行，并返回一个包含每行内容的列表
        data = [line.strip() for line in data] #每一行去除行首和行尾的空白字符（包括空格和换行符），并生成一个新的列表
    return data

# Step 2: 调用机器翻译API
def translate(text, api='youdao'):
    max_retries = 3 #最多尝试3次翻译

    def youdao_translate():
        youdao_key = get_next_youdao_key()
        app_key, app_secret = youdao_key.split(':')
        salt = '12345'
        sign = hashlib.md5((app_key + text + salt + app_secret).encode('utf-8')).hexdigest()
        params = {
            'q': text,
            'from': 'en',
            'to': 'zh',
            'appKey': app_key,
            'salt': salt,
            'sign': sign
        }
        url = "https://openapi.youdao.com/api"
        #发送POST请求到有道API，并解析响应为JSON格式
        response = requests.post(url, data=params)
        result = response.json()
        #如果响应中包含translation，则返回翻译结果，否则打印错误信息并返回"Error"
        if 'translation' in result:
            return result['translation'][0]
        else:
            error_msg = result.get('error_msg', 'Unknown error')
            print(f"Youdao API error: {error_msg}")
            return "Error"

    def baidu_translate():
        baidu_key = get_next_baidu_key()
        appid, api_key = baidu_key.split(':')
        salt = '12345'
        sign = hashlib.md5((appid + text + salt + api_key).encode('utf-8')).hexdigest()
        params = {
            'q': text,
            'from': 'en',
            'to': 'zh',
            'appid': appid,
            'salt': salt,
            'sign': sign
        }
        url = "https://fanyi-api.baidu.com/api/trans/vip/translate"
        response = requests.get(url, params=params)
        result = response.json()
        if 'trans_result' in result:
            return result['trans_result'][0]['dst']
        else:
            error_msg = result.get('error_msg', 'Unknown error')
            print(f"Baidu API error: {error_msg}")
            return "Error"

    def tencent_translate():
        try:
          tencent_key = get_next_tencent_key()
          secret_id, secret_key = tencent_key.split(':')
          cred = credential.Credential(secret_id, secret_key)#创建凭证
          httpProfile = HttpProfile()#设置终端
          httpProfile.endpoint = "tmt.tencentcloudapi.com"

          clientProfile = ClientProfile()
          clientProfile.httpProfile = httpProfile

          client = tmt_client.TmtClient(cred, "ap-guangzhou", clientProfile)#创建腾讯翻译API的客户端

          req = models.TextTranslateRequest()
          params = {
              "SourceText": text,
              "Source": "en",
              "Target": "zh",
              "ProjectId": 0
          }
          req.from_json_string(json.dumps(params))

          resp = client.TextTranslate(req)
          result = json.loads(resp.to_json_string())
          return result['TargetText']
        except TencentCloudSDKException as err:
          print(f"Tencent API error: {err}")
          return "Error"

    def xfyun_translate():
        xfyun_key = get_next_xfyun_key()
        app_id, api_key, api_secret = xfyun_key.split(':')

        host = "itrans.xfyun.cn"
        request_uri = "/v2/its"
        url = "https://" + host + request_uri
        http_method = "POST"
        algorithm = "hmac-sha256"
        http_proto = "HTTP/1.1"
        cur_time_utc = datetime.datetime.utcnow()
        date = cur_time_utc.strftime('%a, %d %b %Y %H:%M:%S GMT')

        def hashlib_256(res):
            m = hashlib.sha256(bytes(res.encode('utf-8'))).digest()
            result = "SHA-256=" + base64.b64encode(m).decode('utf-8')
            return result

        def generate_signature(digest):
            signature_str = "host: " + host + "\n"
            signature_str += "date: " + date + "\n"
            signature_str += http_method + " " + request_uri + " " + http_proto + "\n"
            signature_str += "digest: " + digest
            signature = hmac.new(bytes(api_secret.encode('utf-8')),
                                 bytes(signature_str.encode('utf-8')),
                                 digestmod=hashlib.sha256).digest()
            result = base64.b64encode(signature)
            return result.decode('utf-8')

        digest = hashlib_256(text)
        sign = generate_signature(digest)
        auth_header = f'api_key="{api_key}", algorithm="{algorithm}", headers="host date request-line digest", signature="{sign}"'

        headers = {
            "Content-Type": "application/json",
            "Accept": "application/json",
            "Method": "POST",
            "Host": host,
            "Date": date,
            "Digest": digest,
            "Authorization": auth_header
        }

        business_args = {
            "from": "en",
            "to": "cn",
        }

        content = base64.b64encode(text.encode('utf-8')).decode('utf-8')
        postdata = {
            "common": {"app_id": app_id},
            "business": business_args,
            "data": {"text": content}
        }
        body = json.dumps(postdata)

        response = requests.post(url, data=body, headers=headers, timeout=8)

        result = response.json()
        if 'code' in result and result['code'] == 0:
            return result['data']['result']['trans_result']['dst']
        else:
            error_msg = result.get('message', 'Unknown error')
            print(f"xfyun API error: {error_msg}")
            return None

    def chatgpt_translate():
        api_key = get_next_openai_key()
        os.environ["OPENAI_API_KEY"] = api_key
        os.environ["OPENAI_BASE_URL"] = "https://gtapi.xiaoerchaoren.com:8932/v1"

        client = OpenAI()
        try:
            response = client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "You are a helpful assistant."},
                    {"role": "user", "content": f"Translate the following text from English to Chinese: {text}"}
                ]
            )
            #print(f"ChatGPT Response: {response}")  # 打印响应以调试
            if response and response.choices:
                return response.choices[0].message.content.strip()
            else:
                print("ChatGPT API error: No valid response choices")
                return "Error"
        except requests.exceptions.HTTPError as e:
            if e.response.status_code == 429:
                print("Rate limit exceeded. Retrying...")
                time.sleep(2)  # 增加重试等待时间
                return chatgpt_translate()
            print(f"ChatGPT API exception: {str(e)}")
            return "Error"

    def zhipu_translate():
        zhipu_key = get_next_zhipu_key()
        os.environ["ZHIPU_API_KEY"] = zhipu_key
        client = ZhipuAI(api_key=zhipu_key)
        try:
            response = client.chat.completions.create(
                model="glm-4",
                messages=[
                    {"role": "system", "content": "You are a helpful assistant."},
                    {"role": "user", "content": f"Translate the following English text to Chinese: {text}"}
                ],
                stream=True,
            )
            translation_result = ""
            for chunk in response:
                translation_result += chunk.choices[0].delta.content
            return translation_result.strip()
        except Exception as e:
            print(f"Zhipu API exception: {str(e)}")
            return "Error"
    #定义字典，将API名称映射到相应的翻译函数，更加简洁可读，避免冗长的条件句，轻松管理多个api
    api_translate_function = {
        'youdao': youdao_translate,
        'baidu': baidu_translate,
        'tencent': tencent_translate,
        'xfyun': xfyun_translate,
        'chatgpt': chatgpt_translate,
        'zhipu': zhipu_translate
    }
    #进行最多 max_retries 次的重试，调用选定的翻译函数，获取翻译结果
    for attempt in range(max_retries):
        try:
            translation = api_translate_function[api]()
            if translation != "Error":
                return translation
        except Exception as e:
            print(f"Error in translation attempt {attempt + 1} for {api}: {e}")
        time.sleep(1)  # 等待1秒后重试，避免触发速率限制
    return "Error"

# Step 3: 将翻译结果输入到Excel表格
def write_to_excel(translations, output_file):
    df_total = pd.DataFrame(translations) #将translations转换为一个pandas数据框df_total
    #创建新的数据框，分别保留机器翻译的结果；copy()确保创建的是独立的副本，防止修改原数据框时影响其他数据
    df_baidu = df_total[['Source', 'Reference', 'Baidu']].copy()
    df_youdao = df_total[['Source', 'Reference', 'Youdao']].copy()
    df_tencent = df_total[['Source', 'Reference', 'Tencent']].copy()
    df_xfyun = df_total[['Source', 'Reference', 'Xfyun']].copy()
    df_chatgpt = df_total[['Source', 'Reference', 'ChatGPT']].copy()
    df_zhipu = df_total[['Source', 'Reference', 'Zhipu']].copy()
    #创建Excel写入器writer，将不同的数据框写入到同一个Excel文件的不同工作表中
    with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
        df_total.to_excel(writer, sheet_name='Total', index=False)
        df_baidu.to_excel(writer, sheet_name='Baidu', index=False)
        df_youdao.to_excel(writer, sheet_name='Youdao', index=False)
        df_tencent.to_excel(writer, sheet_name='Tencent', index=False)
        df_xfyun.to_excel(writer, sheet_name='Xfyun', index=False)
        df_chatgpt.to_excel(writer, sheet_name='ChatGPT', index=False)
        df_zhipu.to_excel(writer, sheet_name='Zhipu', index=False)

# Step 4.1: 使用LangChain+GPT-4进行质量评估
# 读取翻译数据的函数
def read_translation_data(file_path):
    df = pd.read_excel(file_path)
    source_texts = df['Source'].tolist()
    reference_translation = df['Reference'].tolist()
    mt_translations = {
        'Baidu': df['Baidu'].tolist(),
        'Youdao': df['Youdao'].tolist(),
        'Tencent': df['Tencent'].tolist(),
        'Xfyun': df['Xfyun'].tolist(),
        'ChatGPT': df['ChatGPT'].tolist(),
        'Zhipu': df['Zhipu'].tolist()
    }
    scores_with_ref = {key: [] for key in mt_translations.keys()}
    return df, source_texts, reference_translation, mt_translations, scores_with_ref

# 清理文本函数,移除文本中的特殊字符，例如换行符和多余的空白
def clean_text(text):
    return text.replace('\n', ' ').replace('\r', '').strip()

# 处理每一行数据
def evaluate_row(row, key, reference=True):
    input_data = {
        "input": f"原文: {row['Source']};中文: {row[key]};"
    }
    if reference:
        input_data["input"] += f"参考译文: {row['Reference']}"

    max_retries = 3
    local_chain = chain
    for attempt in range(max_retries):
        try:
            api_key = get_next_openai_key()
            # print(f"Using API key: {api_key}")
            os.environ["OPENAI_API_KEY"] = api_key
            """ if api_key not in used_openai_keys:
                used_openai_keys.append(api_key) """
            response = local_chain.invoke(input_data)
            response_dict = response.tool_calls
            # 清理模型输出中的多余字符
            for error in response_dict:
                errors = error['args']['errors']
                for err in errors:
                    err['source_chunk'] = clean_text(err['source_chunk'])
                    err['target_chunk'] = clean_text(err['target_chunk'])
                    err['reference_chunk'] = clean_text(err['reference_chunk'])
                    err['explanation'] = clean_text(err['explanation'])  # 清理 explanation 字段

                error_types = [err['error_type'] for err in errors]
                severities = [err.get('severity', 'None') for err in errors]
                explanations = [err['explanation'] for err in errors]  # 获取 explanation 字段
                scores = error['args']['scores']
                return error_types, severities, explanations, scores  # 返回 explanations
        except Exception as e:
            print(f"Error in evaluation attempt {attempt + 1}: {e}")
            if "该令牌额度已用尽" in str(e):
                print("API密钥额度已用尽，正在尝试切换到下一个密钥...")
                time.sleep(2)
                api_key = get_next_openai_key()
                os.environ["OPENAI_API_KEY"] = api_key
                llm.api_key = api_key
                llm_with_tools = llm.bind_tools(tools, tool_choice="TranslationErrorReport")
                local_chain = prompt | llm_with_tools
            else:
                break
    return [], [], [], 0  # 返回空的 explanations

# 添加MQM评估结果到Excel
def add_mqm_evaluation_to_excel(input_file, output_file):
    df, source_texts, reference_translation, mt_translations, scores_with_ref = read_translation_data(input_file)

    # 为存储评估结果创建数据帧
    sub_dfs = {key: df[['Source', 'Reference', key]].copy() for key in mt_translations.keys()}

    # 在主数据帧中添加评估结果列
    for key in mt_translations.keys():
        df[f"{key}_scores"] = 0.0
        sub_dfs[key][f"{key}_error_type"] = ""
        sub_dfs[key][f"{key}_severity"] = ""
        sub_dfs[key][f"{key}_explanation"] = ""  # 新增 explanation 列
        sub_dfs[key][f"{key}_scores"] = 0.0

    # 处理每一行数据
    def process_row(index, row):
        for key in mt_translations.keys():
            error_types, severities, explanations, scores = evaluate_row(row, key, True)  # 获取 explanations
            if error_types is None:
                error_types = []
            if severities is None:
                severities = []
            if explanations is None:
                explanations = []
            if scores is None:
                scores = 0

            # 添加到主数据帧
            df.at[index, f"{key}_scores"] = scores

            # 添加到各个子数据帧
            sub_dfs[key].at[index, f"{key}_error_type"] = ", ".join(error_types)
            sub_dfs[key].at[index, f"{key}_severity"] = ", ".join(map(str, severities))
            sub_dfs[key].at[index, f"{key}_explanation"] = "\n".join(explanations)  # 添加 explanation
            sub_dfs[key].at[index, f"{key}_scores"] = scores

            scores_with_ref[key].append(scores)

    # 使用ThreadPoolExecutor并行处理行数据
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = [executor.submit(process_row, index, row) for index, row in df.iterrows()]
        for future in as_completed(futures):
            try:
                future.result()
            except Exception as e:
                print(f"Error processing row: {e}")

    # 计算每个评分的平均分并添加到数据框中
    df_with_ref = df[['Source', 'Reference']]
    for key in mt_translations.keys():
        df_with_ref[key] = df[key]
        df_with_ref[key + '_score'] = df[f"{key}_scores"]

    # 计算数值列的平均值
    numeric_columns = [col for col in df_with_ref.columns if df_with_ref[col].dtype in ['float64', 'int64']]
    df_with_ref_mean = df_with_ref[numeric_columns].mean().to_frame().T

    # 添加 'Source' 列
    df_with_ref_mean['Source'] = 'Average'
    df_with_ref_mean['Reference'] = ""

    # 确保所有列都包含在 df_with_ref_mean 中
    for col in df_with_ref.columns:
        if col not in df_with_ref_mean.columns:
            df_with_ref_mean[col] = ""

    df_with_ref = pd.concat([df_with_ref, df_with_ref_mean], ignore_index=True)

    # 将所有数据帧写入Excel文件
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_with_ref.to_excel(writer, sheet_name='MQM_reference', index=False)
        for key, sub_df in sub_dfs.items():
            sub_df.to_excel(writer, sheet_name=key, index=False)
    print("MQM翻译质量评估完成并保存到 translations_with_comet_and_mqm_scores.xlsx 文件中。")

# Step 4.2: 使用COMET进行质量评估
# 添加COMET评估结果到Excel
def add_comet_evaluation_to_excel(input_file, output_file):
    df, chinese_texts, reference_translation, mt_translations, scores_with_ref = read_translation_data(input_file)

    # 有参考译文
    def create_dataset_with_ref(src_texts, mt_texts, ref_texts):
        return [{"src": src, "mt": mt, "ref": ref} for src, mt, ref in zip(src_texts, mt_texts, ref_texts)]

    scores_with_ref = {key: [] for key in mt_translations.keys()}
    for key, translations in mt_translations.items():
        data_mt = create_dataset_with_ref(chinese_texts, translations, reference_translation)
        model_output_mt = model_with_ref.predict(data_mt, batch_size=8)
        scores_with_ref[key] = model_output_mt['scores']

    # 无参考译文
    def create_dataset_without_ref(src_texts, mt_texts):
        return [{"src": src, "mt": mt} for src, mt in zip(src_texts, mt_texts)]

    scores_without_ref = {key: [] for key in mt_translations.keys()}
    for key, translations in mt_translations.items():
        data_mt = create_dataset_without_ref(chinese_texts, translations)
        model_output_mt = model_no_ref.predict(data_mt, batch_size=8)
        scores_without_ref[key] = model_output_mt['scores']

    # 将有参考译文的评分结果写入新的表单
    df_with_ref = pd.DataFrame({
        'Source': chinese_texts,
        'Reference': reference_translation
    })
    for key in mt_translations.keys():
        df_with_ref[key] = mt_translations[key]
        df_with_ref[key + '_score'] = scores_with_ref[key]

    # 创建无参考译文的DataFrame
    df_without_ref = pd.DataFrame({
        'Source': chinese_texts,
        'Reference': reference_translation
    })
    for key in mt_translations.keys():
        df_without_ref[key] = mt_translations[key]
        df_without_ref[key + '_score'] = scores_without_ref[key]

    # 计算每个评分的平均分并添加到数据框中
    df_with_ref_mean = df_with_ref.mean(numeric_only=True).to_frame().T
    df_with_ref_mean['Source'] = 'Average'

    df_without_ref_mean = df_without_ref.mean(numeric_only=True).to_frame().T
    df_without_ref_mean['Source'] = 'Average'

    df_with_ref = pd.concat([df_with_ref, df_with_ref_mean], ignore_index=True)
    df_without_ref = pd.concat([df_without_ref, df_without_ref_mean], ignore_index=True)

    # 保存评分结果到新的Excel文件，包含两个表单
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
        df_with_ref.to_excel(writer, sheet_name='comet_reference', index=False)
        df_without_ref.to_excel(writer, sheet_name='comet_no_reference', index=False)

    print("COMET翻译质量评估完成并保存到 translations_with_comet_and_mqm_scores.xlsx 文件中。")

# Step 5: 数据可视化

# 生成并插入折线图的函数
def generate_and_insert_plots(excel_path, sheet_names, output_path):
    # 提取评分数据并生成折线图
    def extract_scores_and_plot(excel_path, sheet_name, output_path):
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
        score_columns = [col for col in df.columns if col.endswith('_score')]
        
        plt.figure(figsize=(10, 6))
        for col in score_columns:
            plt.plot(df.index, df[col], label=col)
        
        plt.xlabel('Reference Index')
        plt.ylabel('Score')
        plt.title(f'Translation Quality Scores - {sheet_name}')
        plt.legend()
        plt.grid(True)
        
        plot_filename = f'{sheet_name}_scores.png'
        plt.savefig(plot_filename)
        plt.close()
        
        # 将折线图插入Excel表单
        wb = load_workbook(excel_path)
        ws = wb[sheet_name]
        
        img = Image(plot_filename)
        img.anchor = 'A15'  # 指定插入图片的位置
        ws.add_image(img)
        
        # 将折线图插入到 "images" 表
        if "images" not in wb.sheetnames:
            wb.create_sheet("images")
        ws_images = wb["images"]
        img_images = Image(plot_filename)
        img_images.width, img_images.height = img.width, img.height  # 保持图片大小一致
        img_images.anchor = f'Q{len(ws_images._images) * 31 + 1}'  # 动态计算插入位置
        ws_images.add_image(img_images)
        
        wb.save(output_path)
        os.remove(plot_filename)  # 删除临时图片文件

    for sheet_name in sheet_names:
        extract_scores_and_plot(excel_path, sheet_name, output_path)
    print("折线图生成并插入到Excel文件中。")
    
# 生成并插入柱状图的函数
def generate_and_insert_bar_charts(excel_path, sheet_names, output_path):
    # 加载Excel文件
    xls = pd.ExcelFile(excel_path)

    # 加载各个分表
    data_frames = {sheet_name: pd.read_excel(excel_path, sheet_name=sheet_name) for sheet_name in sheet_names}

    # 计算每个分表的平均分
    avg_scores = {sheet_name: df.mean(numeric_only=True) for sheet_name, df in data_frames.items()}

    # 定义颜色
    colors = ['skyblue', 'salmon', 'lightgreen', 'orange', 'purple', 'cyan', 'magenta']

    # 定义插入图表的函数
    def insert_chart_to_sheet(sheet, avg_scores, start_row, colors, chart_title):
        fig, ax = plt.subplots()
        ax.bar(avg_scores.index, avg_scores.values, color=colors[:len(avg_scores)])
        ax.set_title(chart_title)
        ax.set_ylabel('Average Score')
        ax.set_xlabel('Model')
        ax.set_xticks(range(len(avg_scores)))
        ax.set_xticklabels(avg_scores.index, rotation=45, ha='right')
        plt.tight_layout()

        # 将图表保存为临时文件
        img_data = BytesIO()
        plt.savefig(img_data, format='png')
        img_data.seek(0)
        img = Image(img_data)
        cell = f'Q{start_row}'
        sheet.add_image(img, cell)
        plt.close(fig)

    # 定义插入综合柱状图的函数
    def insert_combined_chart_to_images(excel_path, ws_images, start_row):
        # 加载Excel文件
        file_path = excel_path
        excel_data = pd.ExcelFile(file_path)

        # 加载相关的工作表到数据框中
        mqm_reference_df = pd.read_excel(file_path, sheet_name='MQM_reference')
        comet_reference_df = pd.read_excel(file_path, sheet_name='comet_reference')
        comet_no_reference_df = pd.read_excel(file_path, sheet_name='comet_no_reference')

        # 只计算数值列的平均分
        mqm_reference_avg = mqm_reference_df.select_dtypes(include='number').mean()
        comet_reference_avg = comet_reference_df.select_dtypes(include='number').mean()
        comet_no_reference_avg = comet_no_reference_df.select_dtypes(include='number').mean()

        # 创建一个新的数据框以合并平均分
        average_scores = pd.DataFrame({
            'Model': ['Baidu', 'Youdao', 'Tencent', 'Xfyun', 'ChatGPT', 'Zhipu'],
            'MQM_reference': [mqm_reference_avg['Baidu_score'], mqm_reference_avg['Youdao_score'], mqm_reference_avg['Tencent_score'], mqm_reference_avg['Xfyun_score'], mqm_reference_avg['ChatGPT_score'], mqm_reference_avg['Zhipu_score']],
            'comet_reference': [comet_reference_avg['Baidu_score'], comet_reference_avg['Youdao_score'], comet_reference_avg['Tencent_score'], comet_reference_avg['Xfyun_score'], comet_reference_avg['ChatGPT_score'], comet_reference_avg['Zhipu_score']],
            'comet_no_reference': [comet_no_reference_avg['Baidu_score'], comet_no_reference_avg['Youdao_score'], comet_no_reference_avg['Tencent_score'], comet_no_reference_avg['Xfyun_score'], comet_no_reference_avg['ChatGPT_score'], comet_no_reference_avg['Zhipu_score']]
        })

        # 创建一个包含每个模型三个柱状的柱状图
        fig, ax = plt.subplots(figsize=(10, 6))

        # 设置柱的位置和宽度
        bar_width = 0.25
        r1 = range(len(average_scores['Model']))
        r2 = [x + bar_width for x in r1]
        r3 = [x + bar_width for x in r2]

        # 创建柱状
        ax.bar(r1, average_scores['MQM_reference'], color='b', width=bar_width, edgecolor='grey', label='MQM_reference')
        ax.bar(r2, average_scores['comet_reference'], color='g', width=bar_width, edgecolor='grey', label='comet_reference')
        ax.bar(r3, average_scores['comet_no_reference'], color='r', width=bar_width, edgecolor='grey', label='comet_no_reference')

        # 添加标签
        ax.set_xlabel('Model', fontweight='bold')
        ax.set_ylabel('Average Score', fontweight='bold')
        ax.set_title('Average Scores by Model and Reference Type')
        ax.set_xticks([r + bar_width for r in range(len(average_scores['Model']))])
        ax.set_xticklabels(average_scores['Model'])

        # 将图例放在图表外部
        ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left')

        # 显示图表
        plt.tight_layout()

        # 将图表保存为字节数组
        img_data = BytesIO()
        plt.savefig(img_data, format='png')
        img_data.seek(0)
        img_bytes = img_data.getvalue()
        img_data.close()

        # 将综合柱状图插入到 "images" 表
        img_images = Image(BytesIO(img_bytes))
        img_images.anchor = f'A{start_row}'  # 动态计算插入位置
        ws_images.add_image(img_images)
        plt.close(fig)

    # 加载工作簿
    wb = load_workbook(excel_path)

    # 插入图表到各自的分表中
    for sheet_name, df in data_frames.items():
        insert_chart_to_sheet(wb[sheet_name], avg_scores[sheet_name], 15, colors, f'{sheet_name} Average Score')

    # 插入综合柱状图到 "images" 表
    if "images" not in wb.sheetnames:
        wb.create_sheet("images")
    ws_images = wb["images"]
    insert_combined_chart_to_images(excel_path, ws_images, (len(ws_images._images)-3) * 30 + 1)

    # 保存工作簿
    wb.save(output_path)
    print("柱状图生成并插入到Excel文件中。")

# 生成并插入饼状图的函数
def generate_and_insert_pie_charts(file_path):
    # 定义获取莫兰迪色系的方法
    def get_morandi_shades(color, n):
        h, l, s = colorsys.rgb_to_hls(*mcolors.to_rgb(color))
        shades = [colorsys.hls_to_rgb(h, l * (1 + i / (2 * n)), s * 0.8) for i in range(n)]
        return shades

    # 定义主色调（莫兰迪色系）
    main_colors = ['#B57C82', '#6E92A3', '#A39EBB', '#A6BAAF', '#B4A29E', '#A2886D', '#5B7493']

    # 创建详细饼图
    def create_detailed_pie_chart(df, title):
        categories = df['Category'].unique()
        category_sizes = df.groupby('Category', sort=False)['Count'].sum().tolist()
        subcategory_labels = df['Subcategory'].tolist()
        subcategory_sizes = df['Count'].tolist()
        fig, ax = plt.subplots(figsize=(8, 6))  # 调整图像大小以适应Excel
        ax.axis('equal')

        # 定义内层颜色
        inner_colors = main_colors[:len(categories)]

        # 定义外层颜色
        outer_colors = []
        for base_color in inner_colors:
            subcategories = df[df['Category'] == categories[inner_colors.index(base_color)]]['Subcategory']
            n = len(subcategories)
            shades = get_morandi_shades(base_color, n)
            outer_colors.extend(shades)

        # 画内层饼图
        wedges, texts, autotexts = ax.pie(category_sizes, radius=1, labels=categories, labeldistance=0.4, autopct='%1.1f%%', pctdistance=0.7, colors=inner_colors)

        # 画外层饼图
        wedges2, texts2, autotexts2 = ax.pie(subcategory_sizes, radius=1.3, labels=subcategory_labels, labeldistance=1.0, autopct='%1.1f%%', pctdistance=0.75, colors=outer_colors, wedgeprops=dict(width=0.4, edgecolor='w'))

        # 调整标签字体大小
        plt.setp(autotexts, size=10, weight="bold", color="white")
        plt.setp(autotexts2, size=8, weight="bold", color="white")
        plt.setp(texts, size=10)
        plt.setp(texts2, size=8)

        plt.title(title, fontsize=16, pad=20)

        # 保存饼状图到内存
        buffer = BytesIO()
        plt.savefig(buffer, format='png')
        plt.close(fig)
        buffer.seek(0)
        return buffer

    # 分类和计数错误
    def classify_and_count_errors(error_series, mqm_typology):
        # 初始化错误计数字典
        error_counts = {cat: {subcat: 0 for subcat in subcats} for cat, subcats in mqm_typology.items()}

        for error_list in error_series:
            if pd.isna(error_list):
                continue
            errors = error_list.split(', ')
            for error in errors:
                for category, subcategories in mqm_typology.items():
                    if error in subcategories:
                        error_counts[category][error] += 1
        return error_counts
        
    # 将错误计数转换为 DataFrame
    def counts_to_df(error_counts):
        data = []
        for category, subcategories in error_counts.items():
            for subcategory, count in subcategories.items():
                if count > 0:
                    data.append((category, subcategory, count))
        return pd.DataFrame(data, columns=['Category', 'Subcategory', 'Count'])

    # 定义 MQM Typology
    mqm_typology = {
        'Accuracy': ['Mistranslation', 'Omission', 'Addition', 'Untranslated', 'Over-translation', 'Under-translation','Do Not Translate'],
        'Audience Appropriateness': ['Specific Culture Reference'],
        'Design and Markup': ['Missing Text', 'Character Formatting'],
        'Fluency': ['Grammar', 'Spelling', 'Punctuation', 'Unintelligible ', 'Character Encoding '],
        'Locale Convention': ['Date Format', 'Currency Format', 'Number Format', 'Measurement Format'],
        'Style': ['Inconsistent Style', 'Inappropriate Style', 'Awkward Style', 'Organizational Style', 'Register', 'Unidiomatic Style'],
        'Terminology': ['Inconsistent Terminology', 'Incorrect Terminology', 'Wrong Term']    
    }

    workbook = load_workbook(file_path)
    sheets = ['Baidu', 'Youdao', 'Tencent', 'Xfyun', 'ChatGPT', 'Zhipu']

    for sheet in sheets:
        data = pd.read_excel(file_path, sheet_name=sheet)
        error_counts = classify_and_count_errors(data[f'{sheet}_error_type'], mqm_typology)
        error_df = counts_to_df(error_counts)

        # 创建饼状图并保存到内存
        pie_img = create_detailed_pie_chart(error_df, f"{sheet} Translation Errors")

        # 插入图像到Excel工作表
        ws = workbook[sheet]
        image = Image(pie_img)
        image.width = 400  # 调整图像宽度
        image.height = 300  # 调整图像高度
        ws.add_image(image, 'A15')  # 插入图像的位置

        # 将饼状图插入到 "images" 表
        if "images" not in workbook.sheetnames:
            workbook.create_sheet("images")
        ws_images = workbook["images"]
        img_images = Image(BytesIO(pie_img.getvalue()))  # 使用新的 BytesIO 对象
        img_images.width, img_images.height = image.width, image.height  # 保持图片大小一致
        img_images.anchor = f'AG{(len(ws_images._images)-4) * 16 + 1}'  # 动态计算插入位置
        ws_images.add_image(img_images)

    # 保存修改后的Excel文件
    workbook.save(file_path)
    print("饼状图生成并插入到Excel文件中。")
#############################################################

def main():
    try: #手动定义输入和输出文件路径
        english_file = 'Alice_en_1.txt'
        chinese_file = 'Alice_zh_1.txt'
        translations_excel = 'translations_Alice.xlsx'
        evaluated_excel = 'translations_with_comet_and_mqm_scores_Alice.xlsx'

        #step1 读取文本文件内容
        english_texts = read_txt(english_file)
        chinese_texts = read_txt(chinese_file)
        assert len(english_texts) == len(chinese_texts), "The number of lines in the Chinese and English files must be the same."#中文和英文文件的行数必须相同，如果不相同则抛出异常

        #step2 翻译
        translations = []
        translation_start_time = time.time() #记录开始翻译的时间点
        #并行翻译
        def translate_parallel(english, chinese):
            translation_baidu = translate(english, api='baidu')
            time.sleep(1)
            translation_youdao = translate(english, api='youdao')
            time.sleep(1)
            translation_tencent = translate(english, api='tencent')
            time.sleep(1)
            translation_xfyun = translate(english, api='xfyun')
            time.sleep(1)
            translation_chatgpt = translate(english, api='chatgpt')
            time.sleep(1)
            translation_zhipu = translate(english, api='zhipu')
            time.sleep(1)

            return {
                'Source': english.strip(),
                'Reference': chinese.strip(),
                'Baidu': translation_baidu,
                'Youdao': translation_youdao,
                'Tencent': translation_tencent,
                'Xfyun': translation_xfyun,
                'ChatGPT': translation_chatgpt,
                'Zhipu': translation_zhipu
            }
        #使用多线程并行处理
        with ThreadPoolExecutor(max_workers=10) as executor:
            futures = [executor.submit(translate_parallel, english, chinese) for english, chinese in zip(english_texts, chinese_texts)]
            for future in as_completed(futures):#迭代完成的任务
                try:
                  translations.append(future.result())#尝试将每个任务的结果添加到 translations 列表中
                except Exception as e:
                    print(f"Error translating text: {e}")
        translation_end_time = time.time() #记录翻译结束的时间点
        translation_duration = translation_end_time - translation_start_time #计算翻译所花费的时间并存储

        #step3 将翻译结果写入到Excel文件
        write_to_excel(translations, translations_excel)
        print(f"Translation time: {translation_duration:.2f} seconds") #打印翻译所花费的时间

        #step4.1 调用MQM评估函数
        mqm_evaluation_start_time = time.time() #记录开始评估的时间点
        add_mqm_evaluation_to_excel(translations_excel, evaluated_excel)
        mqm_evaluation_end_time = time.time() #记录评估结束的时间点
        mqm_evaluation_duration = mqm_evaluation_end_time - mqm_evaluation_start_time #计算评估所花费的时间并存储
        print(f"MQM Evaluation time: {mqm_evaluation_duration:.2f} seconds") #打印评估所花费的时间

        #step4.2 调用COMET评估函数
        comet_evaluation_start_time = time.time() #记录开始评估的时间点
        add_comet_evaluation_to_excel(translations_excel, evaluated_excel)
        comet_evaluation_end_time = time.time() #记录评估结束的时间点
        comet_evaluation_duration = comet_evaluation_end_time - comet_evaluation_start_time #计算评估所花费的时间并存储
        print(f"COMET Evaluation time: {comet_evaluation_duration:.2f} seconds") #打印评估所花费的时间

        #step5 可视化
        # 生成并插入折线图
        generate_and_insert_plots(evaluated_excel,
                     ['MQM_reference', 'comet_reference', 'comet_no_reference'],
                     evaluated_excel)

        # 生成并插入柱状图
        generate_and_insert_bar_charts(evaluated_excel,
                        ['MQM_reference', 'comet_reference', 'comet_no_reference'],
                        evaluated_excel)

        # 生成并插入饼图
        generate_and_insert_pie_charts(evaluated_excel)


    #捕捉并打印整个过程中可能发生的任何异常
    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == '__main__':
    main()
